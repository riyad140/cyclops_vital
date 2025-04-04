# -*- coding: utf-8 -*-
"""
Created on Thu Sep 21 13:17:16 2023

@author: imrul
"""

import openpyxl
import pandas as pd
import numpy as np
import os
from openpyxl.utils.dataframe import dataframe_to_rows
#%%

data = np.arange(10)

df = pd.DataFrame(data)

tiffPath = r"G:\Shared drives\Engineering\Optics_Sensing_Imaging\CC-Cyclops\Imrul_python_script_ISP\image_repo\CYC7_A12\20230911\FAT_ronchi80_Cyc7-A12_FAT7A_glass_\analysis_final_img_blank_BF_fov1\Contrast_heat_map_img_blank_BF_fov1.png"

#%% xls file writer to generate report and image
# resultPath = r'G:\Shared drives\Engineering\Optics_Sensing_Imaging\CC-Cyclops\Imrul_python_script_ISP\image_repo\CYC7_A12\rdb'
# filename='report2.xlsx'
# excelFileName=os.path.join(resultPath,filename)

# wb = openpyxl.load_workbook(filename = excelFileName)
# # sheet_ranges = wb['range names']

# ws1 = wb.create_sheet("Mysheet")

# for r in openpyxl.utils.dataframe.dataframe_to_rows(df):
#     ws1.append(r)
    
# for r in openpyxl.utils.dataframe.dataframe_to_rows(df):
#     ws1.append(r)    
    
# wb.save(excelFileName)


# #%%
# wb = openpyxl.load_workbook(filename = excelFileName)
# ws=wb["Mysheet"]
# img = openpyxl.drawing.image.Image(tiffPath)
# ws.add_image(img,'F5')
# wb.save(excelFileName)


#%%

def create_excel(excelFileName,sheetName = 'TestSheet'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheetName
    wb.save(excelFileName)
    return wb
    
def load_excel(excelFileName, addSheet = True, sheetName = 'NewSheet'):
    wb = openpyxl.load_workbook(filename = excelFileName)
    if addSheet is True:
        ws = wb.create_sheet(sheetName)
    wb.save(excelFileName)
    return wb
    
def df_to_excel(excelFileName,df, sheetName = 'TestSheet'):
    wb = load_excel(excelFileName,addSheet = False, sheetName = sheetName)
    ws = wb[sheetName]
    for r in openpyxl.utils.dataframe.dataframe_to_rows(df):
        ws.append(r)
    wb.save(excelFileName)
    
def image_to_excel(excelFileName,tiffPath, sheetName = 'TestSheet', location = 'A0'):
    img = openpyxl.drawing.image.Image(tiffPath)
    wb = load_excel(excelFileName, addSheet = False, sheetName = sheetName)
    ws = wb[sheetName]
    ws.add_image(img,location)
    wb.save(excelFileName)
    
    
    
#%%
# resultPath = r'G:\Shared drives\Engineering\Optics_Sensing_Imaging\CC-Cyclops\Imrul_python_script_ISP\image_repo\CYC7_A12\rdb'
# filename='report11.xlsx'
# excelFileName=os.path.join(resultPath,filename)
# create_excel(excelFileName)
# # df_to_excel(excelFileName,df,sheetName='dataframe')

# image_to_excel(excelFileName,tiffPath,sheetName='TestSheet',location='A5')

#%%
tiffPath = r"G:\Shared drives\Engineering\Optics_Sensing_Imaging\CC-Cyclops\Imrul_python_script_ISP\image_repo\CYC7_A12\20230911\FAT_ronchi80_Cyc7-A12_FAT7A_glass_\analysis_final_img_blank_BF_fov1\Contrast_heat_map_img_blank_BF_fov1.png"
# fileKey = 'Contrast_heat_map_img_blank'
analysisPath = r'G:\Shared drives\Engineering\Optics_Sensing_Imaging\CC-Cyclops\Imrul_python_script_ISP\image_repo\CYC7_A12\rdb'
filename = 'master_analysis12.xlsx'
excelFileName = os.path.join(analysisPath,filename)
create_excel(excelFileName, sheetName="Ronchi80")
image_to_excel(excelFileName, tiffPath,sheetName="Ronchi80",location='A5')    
    